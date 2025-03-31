#%%
from database import Database

db = Database()
#%%
def get_pcs():
    pcs = db.get_pcs()
    return pcs
# %%
pcs = get_pcs()
pc = pcs[0]

pc_id = pc.id

def get_screen_by_pc(pc_id):
    screens = db.get_screens_by_pc(pc_id)
    return screens

screens = get_screen_by_pc(pc_id)

# %%

infos = {}

infos[pc_id] = []

for screen in screens:
    screen_id = screen.id
    views = db.get_views_by_screen(screen_id)
    # print(views)
    screen_dict = {}
    screen_dict['screen_id'] = screen_id
    screen_dict['views'] = []
    
    for view in views:
        view_id = view.id
        layout = db.get_screen_mappings(screen_id=screen_id, view_id=view_id)
        # print(layout)
        #  slot_row=2, slot_col=1, site_id='10207', camera_id='188181'
        
        view_list = []
        
        for layout_item in layout:
            # print(layout_item)
            slot_row = layout_item.slot_row
            slot_col = layout_item.slot_col
            site_id = layout_item.site_id
            camera_id = layout_item.camera_id
            # print(slot_row, slot_col, site_id, camera_id)
            site = db.get_site_by_id(site_id)
            camera = db.get_camera_by_id(camera_id)
            site_name = site.name
            camera_name = camera.name
            rtsp = camera.rtsp_url
            # print(site_name, camera_name, rtsp)
            view_dict = {
                'slot_row': slot_row,
                'slot_col': slot_col,
                'site_name': site_name,
                'camera_name': camera_name,
                'rtsp': rtsp
            }
            view_list.append(view_dict)
        screen_dict['views'].append(view_list)
    infos[pc_id].append(screen_dict)
    # for view in views:
    #     view_id = view.id
    #     layout = db.get_screen_mappings(screen_id=screen_id, view_id=view_id)
    #     # print(layout)
    #     #  slot_row=2, slot_col=1, site_id='10207', camera_id='188181'
        
    #     for layout_item in layout:
    #         # print(layout_item)
    #         slot_row = layout_item.slot_row
    #         slot_col = layout_item.slot_col
    #         site_id = layout_item.site_id
    #         camera_id = layout_item.camera_id
    #         # print(slot_row, slot_col, site_id, camera_id)
    #         site = db.get_site_by_id(site_id)
    #         camera = db.get_camera_by_id(camera_id)
    #         site_name = site.name
    #         camera_name = camera.name
    #         rtsp = camera.rtsp_url
    #         print(site_name, camera_name, rtsp)
            
    

# %%
import json
import pandas as pd

def dict_to_md_table(camera_dict, output_file="camera_configuration.md"):
    """
    Convert the camera dictionary to a markdown table and save it to a file.
    
    Args:
        camera_dict: Dictionary with PC names as keys and lists of screen data as values
        output_file: Path to save the markdown file
    """
    # Output string for the markdown file
    output = "# Camera Configuration\n\n"
    
    # Process each PC and its screens
    for pc_name, screens in camera_dict.items():
        for screen in screens:
            screen_id = screen.get('screen_id', 'Unknown Screen')
            views = screen.get('views', [])
            
            # Group views by view_id if available, or create a single view group
            views_by_id = {}
            for i, view in enumerate(views):
                # Assuming views are already organized or have a view_id
                view_id = view.get('view_id', f'view_{i//9 + 1}')  # Each view has 9 slots (3x3 grid)
                
                if view_id not in views_by_id:
                    views_by_id[view_id] = []
                views_by_id[view_id].append(view)
            
            # Add PC and screen info
            output += f"## {pc_name} - {screen_id}\n\n"
            
            # Create a table for each view
            for view_id, view_slots in views_by_id.items():
                output += f"### {view_id}\n\n"
                output += "| Position | Column 1 | Column 2 | Column 3 |\n"
                output += "|----------|----------|----------|----------|\n"
                
                # Create a 3x3 grid to organize the slots
                grid = [[None for _ in range(3)] for _ in range(3)]
                
                # Fill the grid with the view slots
                for slot in view_slots:
                    row = slot.get('slot_row', 0) - 1  # Adjust to 0-based indexing
                    col = slot.get('slot_col', 0) - 1  # Adjust to 0-based indexing
                    
                    # Ensure row and col are within bounds
                    if 0 <= row < 3 and 0 <= col < 3:
                        grid[row][col] = slot
                
                # Generate table rows for the grid
                for row_idx, row in enumerate(grid):
                    row_str = f"| **Row {row_idx + 1}** "
                    
                    for col_idx, slot in enumerate(row):
                        if slot:
                            cell_content = (
                                f"**Slot {slot.get('slot_row', '-')}-{slot.get('slot_col', '-')}**<br>"
                                f"Site: {slot.get('site_name', 'N/A')}<br>"
                                f"Camera: {slot.get('camera_name', 'N/A')}<br>"
                                f"RTSP: {slot.get('rtsp', 'N/A')}"
                            )
                        else:
                            cell_content = "Empty"
                        
                        row_str += f"| {cell_content} "
                    
                    row_str += "|\n"
                    output += row_str
                
                output += "\n\n"
    
    # Write to file
    with open(output_file, 'w') as f:
        f.write(output)
    
    print(f"Table saved to {output_file}")
    return output

# Example usage
dict_to_md_table(infos)
# %%
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

def dict_to_excel(camera_dict, output_file="./outputs/camera_configuration.xlsx"):
    """
    Convert the camera dictionary to an Excel workbook with a sheet for each screen_id.
    
    Args:
        camera_dict: Dictionary with PC names as keys and lists of screen data as values
        output_file: Path to save the Excel file
    """
    # Create a workbook
    wb = Workbook()
    # Remove the default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Define styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    centered_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
     
    # Process each PC and its screens
    for pc_name, screens in camera_dict.items():
        for screen in screens:
            screen_id = screen.get('screen_id', f'Unknown_Screen_{len(wb.sheetnames) + 1}')
            views_list = screen.get('views', [])
            
            # Create a new sheet for this screen
            sheet_name = screen_id[-31:] if len(screen_id) > 31 else screen_id  # Excel limit is 31 chars
            ws = wb.create_sheet(sheet_name)
            
            # Add title and PC info
            ws['A1'] = f"Camera Configuration: {pc_name} - {screen_id}"
            ws['A1'].font = Font(bold=True, size=14)
            ws.merge_cells('A1:G1')
            ws['A1'].alignment = centered_align
            
            # Current row for adding content
            current_row = 3
            
            # Process each view (each item in the views list is a complete view)
            for view_idx, view_slots in enumerate(views_list, 1):
                # Add view header
                view_id = f"view_{view_idx}"
                ws[f'A{current_row}'] = f"View: {view_id}"
                ws[f'A{current_row}'].font = Font(bold=True, size=12)
                ws.merge_cells(f'A{current_row}:G{current_row}')
                ws[f'A{current_row}'].alignment = centered_align
                
                current_row += 1
                
                # Create table headers
                headers = ['Position', 'Column 1', 'Column 2', 'Column 3']
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.value = header
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = centered_align
                    cell.border = thin_border

                # Set column widths
                ws.column_dimensions['A'].width = 10
                for col in ['B', 'C', 'D']:
                    ws.column_dimensions[col].width = 50
                
                current_row += 1
                
                # Create a 3x3 grid to organize the slots
                grid = [[None for _ in range(3)] for _ in range(3)]
                
                # Fill the grid with the view slots
                for slot in view_slots:
                    row = slot.get('slot_row', 0) - 1  # Adjust to 0-based indexing
                    col = slot.get('slot_col', 0) - 1  # Adjust to 0-based indexing
                    
                    # Ensure row and col are within bounds
                    if 0 <= row < 3 and 0 <= col < 3:
                        grid[row][col] = slot
                
                # Generate table rows for the grid
                for row_idx, row_data in enumerate(grid):
                    # Add row position
                    pos_cell = ws.cell(row=current_row, column=1)
                    pos_cell.value = f"Row {row_idx + 1}"
                    pos_cell.font = Font(bold=True)
                    pos_cell.alignment = centered_align
                    pos_cell.border = thin_border
                    
                    # Add slot data
                    for col_idx, slot in enumerate(row_data):
                        cell = ws.cell(row=current_row, column=col_idx + 2)
                        
                        if slot:
                            cell_content = (
                                f"Slot {slot.get('slot_row', '-')}-{slot.get('slot_col', '-')}\n"
                                f"Site: {slot.get('site_name', 'N/A')}\n"
                                f"Camera: {slot.get('camera_name', 'N/A')}\n"
                                f"RTSP: {slot.get('rtsp', 'N/A')}"
                            )
                        else:
                            cell_content = "Empty"
                        
                        cell.value = cell_content
                        cell.alignment = Alignment(vertical='top', wrap_text=True)
                        cell.border = thin_border
                    
                    # Set row height
                    ws.row_dimensions[current_row].height = 90
                    current_row += 1
                
                # Add spacing between views
                current_row += 2
    
    # Save the workbook
    wb.save(output_file)
    print(f"Excel file saved to {output_file}")
    
    
dict_to_excel(infos)
# %%
 